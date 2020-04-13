# !/usr/bin/python3
# -*- coding: utf-8 -*-
# @Time     : 2020/1/30 15:35
# @Author   : zky_wind
# @Email    : 475854688@qq.com
# @File     : excel_handler.py
# @Software : PyCharm
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class ExcelHandler:
    """封装一个excel的处理器，主要用来读取excel的数据，和写入数据"""
    def __init__(self, filename):
        self.filename = filename

    def open_sheet(self, sheet_name) -> Worksheet:
        """打开一个excel中的sheet表单"""
        wb = openpyxl.load_workbook(self.filename)
        wb.close()
        return wb[sheet_name]

    def get_title(self, sheet_name):
        """获取表单的表头，并返回一个列表"""
        wb = openpyxl.load_workbook(self.filename)
        sheet = self.open_sheet(sheet_name)
        row = sheet[1]
        title_list = []
        for cell in row:
            title_list.append(cell.value)
        return title_list

    def get_data(self, sheet_name):
        """获取表单中的所有数据，与表头形成字典，返回一个嵌套字典的列表"""
        wb = openpyxl.load_workbook(self.filename)
        sheet = self.open_sheet(sheet_name)
        rows = list(sheet.rows)
        data_list = []
        for row in rows[1:]:
            row_list = []
            for cell in row:
                row_list.append(cell.value)
            row_dict = dict(zip(self.get_title(sheet_name), row_list))
            data_list.append(row_dict)
        return data_list

    @staticmethod
    def change_cell(filename, sheet_name, row, column, value):
        """向excel表单的单元格修改或写入数据"""
        wb = openpyxl.load_workbook(filename)
        sheet = wb[sheet_name]
        sheet.cell(row, column).value = value
        wb.save(filename)
        wb.close()


if __name__ == '__main__':
    excel = ExcelHandler(r"C:\Users\86187\Desktop\test_cases.xlsx")
    print(excel.get_data("case_login"))
